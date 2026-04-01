import os
import hmac
import hashlib
import time
import requests
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VENDOR_ID   = os.environ["COUPANG_VENDOR_ID"]
ACCESS_KEY  = os.environ["COUPANG_ACCESS_KEY"]
SECRET_KEY  = os.environ["COUPANG_SECRET_KEY"]
BOT_TOKEN   = os.environ["TELEGRAM_BOT_TOKEN"]
CHAT_ID     = os.environ["TELEGRAM_CHAT_ID"]

BASE_URL = "https://api-gateway.coupang.com"


def make_signature(method, path, query=""):
    datetime_str = datetime.utcnow().strftime("%y%m%dT%H%M%SZ")
    message = datetime_str + method + path + query
    signature = hmac.new(
        SECRET_KEY.encode("utf-8"),
        message.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()
    return {
        "Authorization": f"CEA algorithm=HmacSHA256, access-key={ACCESS_KEY}, signed-date={datetime_str}, signature={signature}",
        "Content-Type": "application/json;charset=UTF-8"
    }


def get_orders(created_at_from, created_at_to, status="INSTRUCT"):
    path = f"/v2/providers/openapi/apis/api/v4/vendors/{VENDOR_ID}/ordersheets"
    query = f"createdAtFrom={created_at_from}&createdAtTo={created_at_to}&status={status}&maxPerPage=100"
    headers = make_signature("GET", path, query)
    res = requests.get(BASE_URL + path + "?" + query, headers=headers, timeout=30)
    if not res.ok:
        print(f"[주문 조회 오류] {res.status_code} {res.text}")
        return []
    data = res.json()
    return data.get("data", {}).get("content", [])


def get_cancels(created_at_from, created_at_to):
    path = f"/v2/providers/seller_api/apis/api/v1/vendors/{VENDOR_ID}/returnRequests"
    query = f"createdAtFrom={created_at_from}&createdAtTo={created_at_to}&maxPerPage=100"
    headers = make_signature("GET", path, query)
    res = requests.get(BASE_URL + path + "?" + query, headers=headers, timeout=30)
    if not res.ok:
        print(f"[환불 조회 오류] {res.status_code} {res.text}")
        return []
    data = res.json()
    return data.get("data", {}).get("content", [])


def parse_orders(orders):
    total_qty    = 0
    total_sales  = 0
    detail_rows  = []

    for order in orders:
        for item in order.get("orderItems", []):
            qty        = item.get("shippingCount", 1)
            unit_price = item.get("orderPrice", 0)
            sales      = qty * unit_price
            total_qty   += qty
            total_sales += sales
            detail_rows.append({
                "주문번호": order.get("orderId", ""),
                "상품명":   item.get("productName", ""),
                "수량":     qty,
                "단가":     unit_price,
                "금액":     sales,
                "주문상태": item.get("status", ""),
            })

    return total_qty, total_sales, detail_rows


def parse_cancels(cancels):
    total_qty    = 0
    total_amount = 0
    detail_rows  = []

    for cancel in cancels:
        for item in cancel.get("returnItems", []):
            qty    = item.get("returnCount", 1)
            amount = item.get("refundAmount", 0)
            total_qty    += qty
            total_amount += amount
            detail_rows.append({
                "환불번호": cancel.get("returnId", ""),
                "상품명":   item.get("productName", ""),
                "수량":     qty,
                "환불금액": amount,
                "환불사유": cancel.get("returnReason", ""),
            })

    return total_qty, total_amount, detail_rows


def create_excel(today_str, sale_qty, sale_amt, sale_rows, cancel_qty, cancel_amt, cancel_rows):
    wb = Workbook()

    HEADER_FILL   = PatternFill("solid", start_color="2E4057")
    SUMMARY_FILL  = PatternFill("solid", start_color="F0F4F8")
    HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    TITLE_FONT    = Font(name="Arial", bold=True, size=14)
    BODY_FONT     = Font(name="Arial", size=10)
    SUMMARY_FONT  = Font(name="Arial", bold=True, size=11)
    CENTER        = Alignment(horizontal="center", vertical="center")
    LEFT          = Alignment(horizontal="left",   vertical="center")
    THIN          = Side(border_style="thin", color="CCCCCC")
    BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def style_header(cell, text):
        cell.value     = text
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER

    def style_body(cell, value, num_format=None, align=CENTER):
        cell.value     = value
        cell.font      = BODY_FONT
        cell.alignment = align
        cell.border    = BORDER
        if num_format:
            cell.number_format = num_format

    # ---- 요약 시트 ----
    ws1 = wb.active
    ws1.title = "일별 요약"
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 22

    ws1["A1"] = f"쿠팡 일일 판매 리포트 ({today_str})"
    ws1["A1"].font      = TITLE_FONT
    ws1["A1"].alignment = LEFT
    ws1.merge_cells("A1:B1")
    ws1.row_dimensions[1].height = 30

    ws1["A2"] = ""
    headers_summary = [("항목", "수치")]
    summary_data = [
        ("판매 수량 (건)",     sale_qty),
        ("판매 매출 (원)",     sale_amt),
        ("환불 수량 (건)",     cancel_qty),
        ("환불 금액 (원)",     cancel_amt),
        ("실 매출 (원)",       sale_amt - cancel_amt),
    ]

    row = 3
    style_header(ws1.cell(row, 1), "항목")
    style_header(ws1.cell(row, 2), "수치")
    ws1.row_dimensions[row].height = 22

    for i, (label, value) in enumerate(summary_data, start=row + 1):
        ws1.cell(i, 1).value     = label
        ws1.cell(i, 1).font      = SUMMARY_FONT
        ws1.cell(i, 1).fill      = SUMMARY_FILL
        ws1.cell(i, 1).alignment = LEFT
        ws1.cell(i, 1).border    = BORDER
        ws1.cell(i, 2).value     = value
        ws1.cell(i, 2).font      = SUMMARY_FONT
        ws1.cell(i, 2).fill      = SUMMARY_FILL
        ws1.cell(i, 2).alignment = CENTER
        ws1.cell(i, 2).border    = BORDER
        ws1.cell(i, 2).number_format = '#,##0'
        ws1.row_dimensions[i].height = 22

    # ---- 판매 상세 시트 ----
    ws2 = wb.create_sheet("판매 상세")
    sale_headers = ["주문번호", "상품명", "수량", "단가", "금액", "주문상태"]
    col_widths    = [20, 40, 10, 15, 15, 15]
    for col, (h, w) in enumerate(zip(sale_headers, col_widths), 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
        style_header(ws2.cell(1, col), h)
    ws2.row_dimensions[1].height = 22

    for r, row_data in enumerate(sale_rows, start=2):
        style_body(ws2.cell(r, 1), str(row_data["주문번호"]), align=CENTER)
        style_body(ws2.cell(r, 2), row_data["상품명"], align=LEFT)
        style_body(ws2.cell(r, 3), row_data["수량"], '#,##0')
        style_body(ws2.cell(r, 4), row_data["단가"], '#,##0')
        style_body(ws2.cell(r, 5), row_data["금액"], '#,##0')
        style_body(ws2.cell(r, 6), row_data["주문상태"], align=CENTER)
        ws2.row_dimensions[r].height = 20

    if sale_rows:
        total_row = len(sale_rows) + 2
        ws2.cell(total_row, 2).value         = "합계"
        ws2.cell(total_row, 2).font          = Font(name="Arial", bold=True)
        ws2.cell(total_row, 2).border        = BORDER
        ws2.cell(total_row, 3).value         = f"=SUM(C2:C{total_row-1})"
        ws2.cell(total_row, 3).number_format = '#,##0'
        ws2.cell(total_row, 3).font          = Font(name="Arial", bold=True)
        ws2.cell(total_row, 3).border        = BORDER
        ws2.cell(total_row, 5).value         = f"=SUM(E2:E{total_row-1})"
        ws2.cell(total_row, 5).number_format = '#,##0'
        ws2.cell(total_row, 5).font          = Font(name="Arial", bold=True)
        ws2.cell(total_row, 5).border        = BORDER

    # ---- 환불 상세 시트 ----
    ws3 = wb.create_sheet("환불 상세")
    cancel_headers = ["환불번호", "상품명", "수량", "환불금액", "환불사유"]
    col_widths3    = [20, 40, 10, 15, 30]
    for col, (h, w) in enumerate(zip(cancel_headers, col_widths3), 1):
        ws3.column_dimensions[get_column_letter(col)].width = w
        style_header(ws3.cell(1, col), h)
    ws3.row_dimensions[1].height = 22

    for r, row_data in enumerate(cancel_rows, start=2):
        style_body(ws3.cell(r, 1), str(row_data["환불번호"]), align=CENTER)
        style_body(ws3.cell(r, 2), row_data["상품명"], align=LEFT)
        style_body(ws3.cell(r, 3), row_data["수량"], '#,##0')
        style_body(ws3.cell(r, 4), row_data["환불금액"], '#,##0')
        style_body(ws3.cell(r, 5), row_data["환불사유"], align=LEFT)
        ws3.row_dimensions[r].height = 20

    if cancel_rows:
        total_row = len(cancel_rows) + 2
        ws3.cell(total_row, 2).value         = "합계"
        ws3.cell(total_row, 2).font          = Font(name="Arial", bold=True)
        ws3.cell(total_row, 2).border        = BORDER
        ws3.cell(total_row, 3).value         = f"=SUM(C2:C{total_row-1})"
        ws3.cell(total_row, 3).number_format = '#,##0'
        ws3.cell(total_row, 3).font          = Font(name="Arial", bold=True)
        ws3.cell(total_row, 3).border        = BORDER
        ws3.cell(total_row, 4).value         = f"=SUM(D2:D{total_row-1})"
        ws3.cell(total_row, 4).number_format = '#,##0'
        ws3.cell(total_row, 4).font          = Font(name="Arial", bold=True)
        ws3.cell(total_row, 4).border        = BORDER

    filename = f"coupang_{today_str}.xlsx"
    wb.save(filename)
    return filename


def send_telegram_message(text):
    requests.post(
        f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
        data={"chat_id": CHAT_ID, "text": text, "parse_mode": "HTML"}
    )


def send_telegram_file(filepath):
    with open(filepath, "rb") as f:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
            data={"chat_id": CHAT_ID},
            files={"document": f}
        )


if __name__ == "__main__":
    today     = datetime.now()
    yesterday = today - timedelta(days=1)

    date_from = yesterday.strftime("%Y-%m-%dT00:00:00")
    date_to   = yesterday.strftime("%Y-%m-%dT23:59:59")
    today_str = yesterday.strftime("%Y%m%d")
    display   = yesterday.strftime("%Y년 %m월 %d일")

    print(f"[{display}] 데이터 수집 시작")

    orders  = get_orders(date_from, date_to)
    cancels = get_cancels(date_from, date_to)

    sale_qty, sale_amt, sale_rows     = parse_orders(orders)
    cancel_qty, cancel_amt, cancel_rows = parse_cancels(cancels)
    net_sales = sale_amt - cancel_amt

    filename = create_excel(today_str, sale_qty, sale_amt, sale_rows, cancel_qty, cancel_amt, cancel_rows)
    print(f"[엑셀 생성 완료] {filename}")

    summary_msg = (
        f"📦 <b>쿠팡 일일 리포트 ({display})</b>\n\n"
        f"🛒 <b>판매</b>\n"
        f"  수량 : {sale_qty:,}건\n"
        f"  매출 : {sale_amt:,}원\n\n"
        f"↩️ <b>환불</b>\n"
        f"  수량 : {cancel_qty:,}건\n"
        f"  금액 : {cancel_amt:,}원\n\n"
        f"💰 <b>실 매출 : {net_sales:,}원</b>"
    )

    send_telegram_message(summary_msg)
    send_telegram_file(filename)
    print("텔레그램 전송 완료")
